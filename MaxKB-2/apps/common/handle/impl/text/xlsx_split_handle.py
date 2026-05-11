# coding=utf-8
"""
    @project: maxkb
    @Author：虎
    @file： xlsx_parse_qa_handle.py
    @date：2024/5/21 14:59
    @desc:
"""
import io
import traceback
from typing import List

import openpyxl
from openpyxl import load_workbook

from common.handle.base_split_handle import BaseSplitHandle
from common.handle.impl.common_handle import xlsx_embed_cells_images
from common.utils.logger import maxkb_logger

splitter = '\n`-----------------------------------`\n'


def post_cell(image_dict, cell_value):
    image = image_dict.get(cell_value, None)
    if image is not None:
        return f'![](./oss/file/{image.id})'
    return cell_value.replace('\n', '<br>').replace('|', '&#124;')


def row_to_md(row, image_dict):
    return '| ' + ' | '.join(
        [post_cell(image_dict, str(cell.value if cell.value is not None else '')) if cell is not None else '' for cell
         in row]) + ' |\n'


def handle_sheet(file_name, sheet, image_dict, limit: int):
    rows = sheet.rows
    paragraphs = []
    result = {'name': file_name, 'content': paragraphs}
    try:
        title_row_list = next(rows)
        title_md_content = row_to_md(title_row_list, image_dict)
        title_md_content += '| ' + ' | '.join(
            ['---' if cell is not None else '' for cell in title_row_list]) + ' |\n'
    except Exception as e:
        return result
    if len(title_row_list) == 0:
        return result
    result_item_content = ''
    for row in rows:
        next_md_content = row_to_md(row, image_dict)
        next_md_content_len = len(next_md_content)
        result_item_content_len = len(result_item_content)
        if len(result_item_content) == 0:
            result_item_content += title_md_content
            result_item_content += next_md_content
        else:
            if result_item_content_len + next_md_content_len < limit:
                result_item_content += next_md_content
            else:
                paragraphs.append({'content': result_item_content, 'title': ''})
                result_item_content = title_md_content + next_md_content
    if len(result_item_content) > 0:
        paragraphs.append({'content': result_item_content, 'title': ''})
    return result


class XlsxSplitHandle(BaseSplitHandle):
    def fill_merged_cells(self, sheet, image_dict):
        data = []

        # 获取第一行作为标题行
        headers = []
        for idx, cell in enumerate(sheet[1]):
            if cell.value is None:
                headers.append(' ' * (idx + 1))
            else:
                headers.append(cell.value)

        # 从第二行开始遍历每一行
        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for col_idx, cell in enumerate(row):
                cell_value = cell.value

                # 如果单元格为空，并且该单元格在合并单元格内，获取合并单元格的值
                if cell_value is None:
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cell_value = sheet[merged_range.min_row][merged_range.min_col - 1].value
                            break

                image = image_dict.get(cell_value, None)
                if image is not None:
                    cell_value = f'![](./oss/file/{image.id})'

                # 使用标题作为键，单元格的值作为值存入字典
                row_data[headers[col_idx]] = cell_value
            data.append(row_data)

        return data

    def handle(self, file, pattern_list: List, with_filter: bool, limit: int, get_buffer, save_image):
        buffer = get_buffer(file)
        try:
            if type(limit) is str:
                limit = int(limit)
            workbook = openpyxl.load_workbook(io.BytesIO(buffer))
            try:
                image_dict: dict = xlsx_embed_cells_images(io.BytesIO(buffer))
                save_image([item for item in image_dict.values()])
            except Exception as e:
                image_dict = {}
            worksheets = workbook.worksheets
            worksheets_size = len(worksheets)
            return [row for row in
                    [handle_sheet(file.name,
                                  sheet,
                                  image_dict,
                                  limit) if worksheets_size == 1 and sheet.title == 'Sheet1' else handle_sheet(
                        sheet.title, sheet, image_dict, limit) for sheet
                     in worksheets] if row is not None]
        except Exception as e:
            maxkb_logger.error(f"Error processing XLSX file {file.name}: {e}, {traceback.format_exc()}")
            return [{'name': file.name, 'content': []}]

    def get_content(self, file, save_image):
        try:
            # 加载 Excel 文件
            workbook = load_workbook(file)
            try:
                image_dict: dict = xlsx_embed_cells_images(file)
                if len(image_dict) > 0:
                    save_image(image_dict.values())
            except Exception as e:
                maxkb_logger.error(f'Exception: {e}')
                image_dict = {}
            md_tables = ''
            # 遍历所有工作表
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                rows = self.fill_merged_cells(sheet, image_dict)
                if len(rows) == 0:
                    continue

                # 添加 sheet 名称作为标题
                md_tables += f'## {sheetname}\n\n'

                # 提取表头和内容
                headers = [f"{key}" for key, value in rows[0].items()]

                # 构建 Markdown 表格
                md_table = '| ' + ' | '.join(headers) + ' |\n'
                md_table += '| ' + ' | '.join(['---'] * len(headers)) + ' |\n'
                for row in rows:
                    r = [self._escape_cell_content(value) for key, value in row.items()]
                    md_table += '| ' + ' | '.join(r) + ' |\n'

                md_tables += md_table + '\n\n'

            return md_tables
        except Exception as e:
            maxkb_logger.error(f'excel split handle error: {e}')
            return f'error: {e}'

    def _escape_cell_content(self, cell_value):
        """转义单元格内容,避免破坏 Markdown 表格结构"""
        if cell_value is None:
            return ''

        cell_str = str(cell_value)

        # 替换换行符为 <br>
        cell_str = cell_str.replace('\n', '<br>')

        # 转义管道符 | 为 HTML 实体
        cell_str = cell_str.replace('|', '&#124;')

        # 如果内容包含反引号,需要转义
        if '`' in cell_str:
            cell_str = cell_str.replace('`', '&#96;')

        return cell_str

    def support(self, file, get_buffer):
        file_name: str = file.name.lower()
        if file_name.endswith(".xlsx"):
            return True
        return False
